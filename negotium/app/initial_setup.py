"""Initial office setup helpers.

The first-run wizard uses these utilities to read uploaded company files and
prepare a compact, LLM-friendly summary. Parsers intentionally keep the output
small and tabular so the LLM can infer roles/users without receiving entire
documents.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

SENSITIVE_KEYWORDS = (
    "주민등록",
    "계약",
    "급여",
    "연봉",
    "고객",
    "개인정보",
    "재무",
    "인사",
    "employee",
    "salary",
    "customer",
    "contract",
    "finance",
)


@dataclass(frozen=True)
class ParsedSetupFile:
    path: str
    filename: str
    kind: str
    text: str
    rows: list[dict[str, str]] = field(default_factory=list)
    sensitive_hint: bool = False

    def to_prompt_block(self) -> str:
        rows_md = "\n".join(
            "- " + ", ".join(f"{key}={value}" for key, value in row.items() if value)
            for row in self.rows[:30]
        )
        return "\n".join(
            [
                f"### {self.filename} ({self.kind})",
                f"- path: {self.path}",
                f"- sensitive_hint: {self.sensitive_hint}",
                self.text[:4000],
                rows_md,
            ],
        ).strip()


def parse_setup_file(path: Path, *, archive_root: Path | None = None) -> ParsedSetupFile:
    suffix = path.suffix.lower()
    relative = path.relative_to(archive_root).as_posix() if archive_root else path.as_posix()
    if suffix in {".csv", ".tsv"}:
        rows = _read_delimited(path, delimiter="\t" if suffix == ".tsv" else ",")
        text = _rows_to_text(rows)
        return ParsedSetupFile(
            path=relative,
            filename=path.name,
            kind=suffix.lstrip("."),
            text=text,
            rows=rows,
            sensitive_hint=_has_sensitive_hint(path.name, text),
        )
    if suffix == ".xlsx":
        rows = _read_xlsx(path)
        text = _rows_to_text(rows)
        return ParsedSetupFile(
            path=relative,
            filename=path.name,
            kind="xlsx",
            text=text,
            rows=rows,
            sensitive_hint=_has_sensitive_hint(path.name, text),
        )
    text = path.read_text(encoding="utf-8", errors="ignore")[:12000]
    return ParsedSetupFile(
        path=relative,
        filename=path.name,
        kind=suffix.lstrip(".") or "text",
        text=text,
        sensitive_hint=_has_sensitive_hint(path.name, text),
    )


def parse_setup_uploads(
    upload_records: list[dict[str, Any]], *, archive_root: Path
) -> list[ParsedSetupFile]:
    parsed: list[ParsedSetupFile] = []
    for record in upload_records:
        rel = str(record.get("path") or "")
        if not rel:
            continue
        path = archive_root / rel
        if path.exists() and path.is_file():
            parsed.append(parse_setup_file(path, archive_root=archive_root))
    return parsed


def _read_delimited(path: Path, *, delimiter: str) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh, delimiter=delimiter)
        return [
            {str(key or "").strip(): str(value or "").strip() for key, value in row.items()}
            for row in reader
        ]


def _read_xlsx(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise RuntimeError("XLSX parsing requires openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    parsed: list[dict[str, str]] = []
    for row in rows[1:101]:
        parsed.append(
            {
                headers[index] or f"col_{index + 1}": str(value or "").strip()
                for index, value in enumerate(row)
            },
        )
    return parsed


def _rows_to_text(rows: list[dict[str, str]]) -> str:
    if not rows:
        return ""
    headers = list(rows[0].keys())
    rendered = [", ".join(headers)]
    for row in rows[:50]:
        rendered.append(", ".join(row.get(header, "") for header in headers))
    return "\n".join(rendered)


def _has_sensitive_hint(filename: str, text: str) -> bool:
    haystack = f"{filename}\n{text[:3000]}".lower()
    return any(keyword.lower() in haystack for keyword in SENSITIVE_KEYWORDS)
